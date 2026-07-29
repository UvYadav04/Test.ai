import openpyxl
import pandas as pd

MIN_TABLE_ROWS = 2
MIN_TABLE_COLS = 1


def load_sheets(file_path: str) -> dict:
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    try:
        return {name: _sheet_grid(workbook[name]) for name in workbook.sheetnames}
    finally:
        workbook.close()


def _sheet_grid(worksheet) -> list:
    grid = [list(row) for row in worksheet.iter_rows(values_only=True)]
    for merged_range in worksheet.merged_cells.ranges:
        _fill_merged_range(grid, merged_range)
    return grid


def _fill_merged_range(grid: list, merged_range) -> None:
    if merged_range.min_row - 1 >= len(grid):
        return
    value = grid[merged_range.min_row - 1][merged_range.min_col - 1]
    for row in range(merged_range.min_row - 1, min(merged_range.max_row, len(grid))):
        row_values = grid[row]
        for col in range(merged_range.min_col - 1, min(merged_range.max_col, len(row_values))):
            row_values[col] = value


def detect_tables(grid: list) -> list:
    tables = []
    for row_start, row_end in _nonblank_ranges(grid):
        row_block = grid[row_start:row_end]
        columns_view = list(zip(*row_block)) if row_block else []
        for col_start, col_end in _nonblank_ranges(columns_view):
            block = [row[col_start:col_end] for row in row_block]
            table = _table_from_block(block, anchor_row=row_start, anchor_col=col_start)
            if table is not None:
                tables.append(table)
    return tables


def _nonblank_ranges(sequences: list) -> list:
    ranges = []
    start = None
    for i, seq in enumerate(sequences):
        blank = _is_blank_sequence(seq)
        if not blank and start is None:
            start = i
        elif blank and start is not None:
            ranges.append((start, i))
            start = None
    if start is not None:
        ranges.append((start, len(sequences)))
    return ranges


def _is_blank_sequence(seq) -> bool:
    return all(_is_blank_cell(cell) for cell in seq)


def _is_blank_cell(value) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def _table_from_block(block: list, anchor_row: int, anchor_col: int):
    if len(block) < MIN_TABLE_ROWS:
        return None

    header, *data_rows = block
    columns = _column_names(header)
    if len(columns) < MIN_TABLE_COLS:
        return None

    data_rows = [row for row in data_rows if not _is_blank_sequence(row)]
    if not data_rows:
        return None

    dataframe = pd.DataFrame(data_rows, columns=columns)
    return {"dataframe": dataframe, "anchor_row": anchor_row, "anchor_col": anchor_col}


def _column_names(header: list) -> list:
    seen: dict = {}
    names = []
    for i, cell in enumerate(header):
        base = str(cell).strip() if not _is_blank_cell(cell) else f"column_{i + 1}"
        if base in seen:
            seen[base] += 1
            base = f"{base}_{seen[base]}"
        else:
            seen[base] = 0
        names.append(base)
    return names
