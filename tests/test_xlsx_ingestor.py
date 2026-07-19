import openpyxl
import pytest

from ingestion.file_types.xlsx.xlsx_ingestor import XLSXIngestor
from ingestion.storage.local_store import LocalParquetStore


@pytest.fixture
def storage(tmp_path):
    return LocalParquetStore(root_dir=str(tmp_path / "parquet"))


def _write_rows(ws, rows, start_row=1, start_col=1):
    for r, row in enumerate(rows):
        for c, value in enumerate(row):
            ws.cell(row=start_row + r, column=start_col + c, value=value)


@pytest.fixture
def single_table_xlsx(tmp_path):
    path = tmp_path / "people.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "People"
    _write_rows(ws, [
        ["name", "age", "city"],
        ["Alice", 30, "NYC"],
        ["Bob", 25, "LA"],
        ["Carol", 40, "SF"],
    ])
    wb.save(path)
    return str(path)


@pytest.fixture
def multi_sheet_xlsx(tmp_path):
    path = tmp_path / "workbook.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Orders"
    _write_rows(ws1, [
        ["order_id", "amount"],
        [1, 100],
        [2, 250],
    ])
    ws2 = wb.create_sheet("Customers")
    _write_rows(ws2, [
        ["customer_id", "name"],
        [1, "Alice"],
        [2, "Bob"],
    ])
    wb.save(path)
    return str(path)


@pytest.fixture
def stacked_tables_xlsx(tmp_path):
    path = tmp_path / "stacked.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"
    _write_rows(ws, [
        ["region", "revenue"],
        ["East", 1000],
        ["West", 2000],
    ], start_row=1)
    # blank row at row 4 separates the two tables
    _write_rows(ws, [
        ["product", "units"],
        ["Widget", 5],
        ["Gadget", 9],
    ], start_row=5)
    wb.save(path)
    return str(path)


@pytest.fixture
def side_by_side_tables_xlsx(tmp_path):
    path = tmp_path / "side_by_side.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    _write_rows(ws, [
        ["region", "revenue"],
        ["East", 1000],
        ["West", 2000],
    ], start_row=1, start_col=1)
    # blank column at column C separates the two tables
    _write_rows(ws, [
        ["product", "units"],
        ["Widget", 5],
        ["Gadget", 9],
    ], start_row=1, start_col=4)
    wb.save(path)
    return str(path)


@pytest.fixture
def merged_cells_xlsx(tmp_path):
    path = tmp_path / "merged.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"
    _write_rows(ws, [
        ["region", "quarter", "revenue"],
        ["East", "Q1", 1000],
        ["East", "Q2", 1500],
        ["West", "Q1", 900],
        ["West", "Q2", 1200],
    ])
    # A real spreadsheet export merging repeated row labels: openpyxl (like Excel) clears
    # every cell in a merge range except the top-left one, so A3/A5 go back to None here even
    # though we just wrote "East"/"West" into them above.
    ws.merge_cells("A2:A3")
    ws.merge_cells("A4:A5")
    wb.save(path)
    return str(path)


@pytest.fixture
def no_table_xlsx(tmp_path):
    path = tmp_path / "notes.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "just a note, no real table here"
    wb.save(path)
    return str(path)


def test_validate_accepts_well_formed_workbook(single_table_xlsx, storage):
    ingestor = XLSXIngestor(storage=storage)
    assert ingestor.validate(single_table_xlsx) is True


def test_validate_rejects_empty_file(tmp_path, storage):
    path = tmp_path / "empty.xlsx"
    path.write_bytes(b"")
    ingestor = XLSXIngestor(storage=storage)
    assert ingestor.validate(str(path)) is False


def test_validate_rejects_missing_file(storage):
    ingestor = XLSXIngestor(storage=storage)
    assert ingestor.validate("/no/such/file.xlsx") is False


def test_validate_rejects_corrupt_file(tmp_path, storage):
    path = tmp_path / "corrupt.xlsx"
    path.write_bytes(b"not a real xlsx file" * 10)
    ingestor = XLSXIngestor(storage=storage)
    assert ingestor.validate(str(path)) is False


def test_validate_rejects_workbook_with_no_real_table(no_table_xlsx, storage):
    ingestor = XLSXIngestor(storage=storage)
    assert ingestor.validate(no_table_xlsx) is False
    assert ingestor.errors


def test_extract_metadata_reports_sheet_and_table_counts(multi_sheet_xlsx, storage):
    ingestor = XLSXIngestor(storage=storage)
    meta = ingestor.extract_metadata(multi_sheet_xlsx)
    assert meta["sheet_count"] == 2
    assert meta["table_count"] == 2
    assert set(meta["sheets"]) == {"Orders", "Customers"}


def test_ingest_single_table_single_sheet(single_table_xlsx, storage):
    ingestor = XLSXIngestor(storage=storage)
    result = ingestor.ingest(single_table_xlsx, workspace_id="ws1", file_id="file1")

    assert result.status == "success"
    assert len(result.extracted_tables) == 1
    table = result.extracted_tables[0]
    assert table["sheet"] == "People"
    assert table["row_count"] == 3
    assert set(table["columns"]) == {"name", "age", "city"}
    assert storage.exists(table["output_ref"])

    df = storage.read(table["output_ref"])
    assert len(df) == 3


def test_ingest_multi_sheet_workbook_produces_one_table_per_sheet(multi_sheet_xlsx, storage):
    ingestor = XLSXIngestor(storage=storage)
    result = ingestor.ingest(multi_sheet_xlsx, workspace_id="ws1", file_id="file2")

    assert result.status == "success"
    assert len(result.extracted_tables) == 2
    sheets = {t["sheet"] for t in result.extracted_tables}
    assert sheets == {"Orders", "Customers"}
    # single table per sheet -> location is just the sheet name, no "table N" suffix
    locations = {t["location"] for t in result.extracted_tables}
    assert locations == {"Orders", "Customers"}


def test_ingest_detects_stacked_tables_in_one_sheet(stacked_tables_xlsx, storage):
    ingestor = XLSXIngestor(storage=storage)
    result = ingestor.ingest(stacked_tables_xlsx, workspace_id="ws1", file_id="file3")

    assert result.status == "success"
    assert len(result.extracted_tables) == 2
    assert all(t["sheet"] == "Report" for t in result.extracted_tables)
    row_counts = sorted(t["row_count"] for t in result.extracted_tables)
    assert row_counts == [2, 2]


def test_ingest_detects_side_by_side_tables_in_one_sheet(side_by_side_tables_xlsx, storage):
    ingestor = XLSXIngestor(storage=storage)
    result = ingestor.ingest(side_by_side_tables_xlsx, workspace_id="ws1", file_id="file4")

    assert result.status == "success"
    assert len(result.extracted_tables) == 2
    column_sets = [set(t["columns"]) for t in result.extracted_tables]
    assert {"region", "revenue"} in column_sets
    assert {"product", "units"} in column_sets


def test_ingest_forward_fills_merged_cell_values_instead_of_losing_them(merged_cells_xlsx, storage):
    ingestor = XLSXIngestor(storage=storage)
    result = ingestor.ingest(merged_cells_xlsx, workspace_id="ws1", file_id="file5")

    assert result.status == "success"
    assert len(result.extracted_tables) == 1
    table = result.extracted_tables[0]
    assert table["row_count"] == 4

    df = storage.read(table["output_ref"])
    # Without forward-filling the merge, rows 2 and 4 of the "region" column would read back
    # as None instead of repeating "East"/"West".
    assert df["region"].tolist() == ["East", "East", "West", "West"]


def test_ingest_failure_when_storage_missing_returns_failed_status(single_table_xlsx):
    ingestor = XLSXIngestor(storage=None)
    result = ingestor.ingest(single_table_xlsx, workspace_id="ws1", file_id="file6")
    assert result.status == "failed"
    assert result.errors


def test_ingest_failure_when_no_tables_detected(no_table_xlsx, storage):
    ingestor = XLSXIngestor(storage=storage)
    result = ingestor.ingest(no_table_xlsx, workspace_id="ws1", file_id="file7")
    assert result.status == "failed"
    assert result.errors


def test_ingest_file_routes_xlsx_through_xlsx_ingestor(single_table_xlsx, storage):
    from ingestion.manager import IngestionManager
    from tests.fakes import FakeVectorStore

    manager = IngestionManager(storage=storage, vector_store=FakeVectorStore())
    result = manager.ingest_file(single_table_xlsx, workspace_id="ws1", file_id="f1")

    assert result.status == "success"
    assert len(result.extracted_tables) == 1
